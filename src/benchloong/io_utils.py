"""结果写入：按附件模板填充 result1 / result2 / result4。"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import openpyxl

ROOT = Path(__file__).resolve().parents[2]
OUTPUT_DIR = ROOT / "results"
TEMPLATES = {
    "result1.xlsx": ROOT / "附件1.xlsx",
    "result2.xlsx": ROOT / "附件2.xlsx",
    "result4.xlsx": ROOT / "附件4.xlsx",
}


def _round6(v: float) -> float:
    return round(float(v), 6)


def _fill_position_sheet(ws, positions: np.ndarray, times):
    """positions: (T, 224, 2)；行号：点 p 的 x/y 在 2+2p / 3+2p 行，列按表头时间匹配。"""
    time_cols = {}
    for c in range(2, ws.max_column + 1):
        header = ws.cell(1, c).value
        if header is None:
            continue
        time_cols[header] = c
    for k, t in enumerate(times):
        col = time_cols.get(t)
        if col is None:
            raise KeyError(f"模板中找不到时间列 {t}")
        for p in range(positions.shape[1]):
            ws.cell(2 + 2 * p, col, _round6(positions[k, p, 0]))
            ws.cell(3 + 2 * p, col, _round6(positions[k, p, 1]))


def _fill_speed_sheet(ws, speeds: np.ndarray, times):
    """speeds: (T, 224)。"""
    time_cols = {}
    for c in range(2, ws.max_column + 1):
        header = ws.cell(1, c).value
        if header is None:
            continue
        time_cols[header] = c
    for k, t in enumerate(times):
        col = time_cols.get(t)
        if col is None:
            raise KeyError(f"模板中找不到时间列 {t}")
        for p in range(speeds.shape[1]):
            ws.cell(2 + p, col, _round6(speeds[k, p]))


def write_result1(positions: np.ndarray, speeds: np.ndarray,
                 times=None, out_name: str = "result1.xlsx") -> Path:
    """问题 1 / 问题 4 共用：位置 + 速度两张表。"""
    if times is None:
        times = [f"{t} s" for t in range(0, 301)]
    wb = openpyxl.load_workbook(TEMPLATES[out_name])
    _fill_position_sheet(wb["位置"], positions, times)
    _fill_speed_sheet(wb["速度"], speeds, times)
    OUTPUT_DIR.mkdir(exist_ok=True)
    out = OUTPUT_DIR / out_name
    wb.save(out)
    return out


def write_result2(x: np.ndarray, y: np.ndarray, speeds: np.ndarray,
                  out_name: str = "result2.xlsx") -> Path:
    """问题 2：终止时刻位置与速度（Sheet1）。"""
    wb = openpyxl.load_workbook(TEMPLATES[out_name])
    ws = wb["Sheet1"]
    for p in range(speeds.shape[0]):
        ws.cell(2 + p, 2, _round6(x[p]))
        ws.cell(2 + p, 3, _round6(y[p]))
        ws.cell(2 + p, 4, _round6(speeds[p]))
    OUTPUT_DIR.mkdir(exist_ok=True)
    out = OUTPUT_DIR / out_name
    wb.save(out)
    return out
